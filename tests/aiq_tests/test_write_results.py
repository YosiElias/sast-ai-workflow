"""
Unit tests for the write_results tool's core function.
"""

import unittest
from unittest.mock import Mock, patch

from sast_agent_workflow.tools.write_results import write_results, WriteResultsConfig
from dto.SASTWorkflowModels import SASTWorkflowTracker
from dto.Issue import Issue
from dto.LLMResponse import AnalysisResponse
from common.config import Config
from aiq.builder.builder import Builder
from tests.aiq_tests.test_utils import TestUtils


class TestWriteResultsCore(unittest.IsolatedAsyncioTestCase):

    def setUp(self):
        self.mock_config = Mock(spec=Config)
        self.mock_config.WRITE_RESULTS = True
        self.mock_config.HUMAN_VERIFIED_FILE_PATH = None  # Add missing attribute
        self.write_results_config = WriteResultsConfig()
        self.builder = Mock(spec=Builder)

    async def test__write_results__final_issues_writes_to_excel_successfully(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="final_issue_1", issue_type="BUFFER_OVERFLOW"),
            TestUtils.create_sample_issue(issue_id="final_issue_2", issue_type="USE_AFTER_FREE")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Issue is a true positive", "Buffer overflow confirmed"],
            short_justifications="True positive"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        tracker.metrics = {"total_issues": 2, "confusion_matrix": {"true_positives": 1, "true_negatives": 1, "false_positives": 0, "false_negatives": 0}}
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer, \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary') as mock_eval_summary:
            
            mock_summary_data = [
                (issues[0], Mock()), 
                (issues[1], Mock())
            ]
            mock_convert.return_value = mock_summary_data
            mock_eval_summary_instance = Mock()
            mock_eval_summary.return_value = mock_eval_summary_instance
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify conversion called with include_non_final=True by default
            mock_convert.assert_called_once_with(tracker, include_non_final=True)
            
            # Verify Excel writer called with correct parameters - using metrics-based evaluation summary
            mock_excel_writer.assert_called_once()
            call_args = mock_excel_writer.call_args[0]
            self.assertEqual(call_args[0], mock_summary_data)  # summary_data
            self.assertEqual(call_args[2], self.mock_config)  # config
            # Verify evaluation_summary was created from metrics (not the mock)
            eval_summary = call_args[1]
            self.assertEqual(eval_summary.tp, 1)
            self.assertEqual(eval_summary.tn, 1)
            self.assertEqual(eval_summary.fp, 0)
            self.assertEqual(eval_summary.fn, 0)

    async def test__write_results__preserves_all_tracker_data_unchanged(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        original_tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        original_tracker.metrics = {"test_metric": "test_value"}
        original_tracker.iteration_count = 5
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data'), \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file'), \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary'):
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, original_tracker)
            
            # assertion - verify terminal node behavior (no state changes)
            self.assertEqual(result_tracker.issues, original_tracker.issues)
            self.assertEqual(result_tracker.config, original_tracker.config)
            self.assertEqual(result_tracker.metrics, original_tracker.metrics)
            self.assertEqual(result_tracker.iteration_count, original_tracker.iteration_count)

    async def test__write_results__write_results_disabled_skips_writing(self):
        # preparation
        self.mock_config.WRITE_RESULTS = False
        
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer:
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify no writing operations were called
            mock_convert.assert_not_called()
            mock_excel_writer.assert_not_called()

    async def test__write_results__no_config_skips_writing_gracefully(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        tracker.config = None  # Explicitly set to None after creation
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer:
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify no writing operations were called
            mock_convert.assert_not_called()
            mock_excel_writer.assert_not_called()

    async def test__write_results__empty_tracker_handles_gracefully(self):
        # preparation
        empty_tracker = SASTWorkflowTracker(config=self.mock_config, issues={})
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer:
            
            mock_convert.return_value = []  # Empty summary data
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, empty_tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            self.assertEqual(len(result_tracker.issues), 0)
            self.assertEqual(result_tracker.config, self.mock_config)
            
            # Verify conversion was called but no issues to write
            mock_convert.assert_called_once_with(empty_tracker, include_non_final=True)
            
            # Excel writer should still be called but with empty data
            mock_excel_writer.assert_called_once()

    async def test__write_results__no_completed_issues_handles_appropriately(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="non_final_issue_1", issue_type="BUFFER_OVERFLOW"),
            TestUtils.create_sample_issue(issue_id="non_final_issue_2", issue_type="USE_AFTER_FREE")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="FALSE",  # All issues are non-final
            justifications=["Under investigation"],
            short_justifications="Pending"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer:
            
            mock_convert.return_value = []  # No final issues to convert
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify conversion was called with include_non_final=True by default
            mock_convert.assert_called_once_with(tracker, include_non_final=True)
            
            # Excel writer should still be called with empty data
            mock_excel_writer.assert_called_once()

    async def test__write_results__excel_writer_failure_handles_gracefully(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer, \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary'):
            
            mock_convert.return_value = [(issues[0], Mock())]
            mock_excel_writer.side_effect = Exception("Excel writing failed")
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify tracker returned unchanged despite error
            self.assertEqual(result_tracker.issues, tracker.issues)
            self.assertEqual(result_tracker.config, tracker.config)

    async def test__write_results__invalid_metrics_triggers_fallback_evaluation_summary_creation(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        # Set invalid metrics to trigger fallback
        tracker.metrics = {"error": "calculation failed"}
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer, \
             patch('sast_agent_workflow.tools.write_results.get_human_verified_results') as mock_get_ground_truth, \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary') as mock_eval_summary, \
             patch('sast_agent_workflow.tools.write_results._create_mock_evaluation_summary') as mock_create_mock:
            
            mock_convert.return_value = [(issues[0], Mock())]
            mock_get_ground_truth.return_value = None
            mock_eval_summary_instance = Mock()
            mock_eval_summary.return_value = mock_eval_summary_instance
            
            # Make _create_mock_evaluation_summary fail to trigger fallback path
            mock_create_mock.side_effect = Exception("Mock creation failed")
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify fallback EvaluationSummary creation was called
            mock_get_ground_truth.assert_called_once_with(self.mock_config)
            mock_eval_summary.assert_called_once()
            
            # Verify Excel writer was called with fallback evaluation summary
            mock_excel_writer.assert_called_once()
            call_args = mock_excel_writer.call_args[0]
            self.assertEqual(call_args[1], mock_eval_summary_instance)  # evaluation_summary

    async def test__write_results__missing_ml_performance_metrics_triggers_fallback_gracefully(self):
        """
        Test that write_results handles missing ML performance metrics gracefully.
        
        tracker.metrics contains ML performance metrics like accuracy, precision, recall, 
        f1_score, and confusion_matrix (NOT ragas metrics like answer_relevancy).
        When these ML metrics are None, the function should attempt fallback 
        EvaluationSummary creation and handle failures gracefully.
        """
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        # Set ML performance metrics to None to trigger fallback evaluation summary creation
        tracker.metrics = None
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer, \
             patch('sast_agent_workflow.tools.write_results.get_human_verified_results') as mock_get_ground_truth, \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary') as mock_eval_summary:
            
            mock_convert.return_value = [(issues[0], Mock())]
            # Make both fallback methods fail
            mock_get_ground_truth.side_effect = Exception("Ground truth failed")
            mock_eval_summary.side_effect = Exception("EvaluationSummary failed")
            
            # testing
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify Excel writer was called with None evaluation_summary
            mock_excel_writer.assert_called_once()
            call_args = mock_excel_writer.call_args[0]
            self.assertIsNone(call_args[1])  # evaluation_summary should be None

    async def test__write_results__evaluation_summary_creation_handles_all_error_paths(self):
        # preparation
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue", issue_type="BUFFER_OVERFLOW")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        # Set None metrics to trigger initial fallback path
        tracker.metrics = None
        
        with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert, \
             patch('sast_agent_workflow.tools.write_results.write_to_excel_file') as mock_excel_writer, \
             patch('sast_agent_workflow.tools.write_results.get_human_verified_results') as mock_get_ground_truth, \
             patch('sast_agent_workflow.tools.write_results.EvaluationSummary') as mock_eval_summary:
            
            mock_convert.return_value = [(issues[0], Mock())]
            mock_get_ground_truth.return_value = None
            mock_eval_summary_instance = Mock()
            mock_eval_summary.return_value = mock_eval_summary_instance
            
            # testing - this should trigger the direct EvaluationSummary creation path
            result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
            
            # assertion
            self.assertIsInstance(result_tracker, SASTWorkflowTracker)
            
            # Verify direct EvaluationSummary creation path was taken
            mock_get_ground_truth.assert_called_once_with(self.mock_config)
            mock_eval_summary.assert_called_once()
            
            # Verify Excel writer was called with the evaluation summary
            mock_excel_writer.assert_called_once()

    async def test__write_results__evaluation_summary_integration_with_real_excel_writer(self):
        # preparation
        import tempfile
        import os
        from unittest.mock import patch
        
        issues = [
            TestUtils.create_sample_issue(issue_id="test_issue_1", issue_type="BUFFER_OVERFLOW"),
            TestUtils.create_sample_issue(issue_id="test_issue_2", issue_type="USE_AFTER_FREE")
        ]
        
        per_issue_data = TestUtils.create_sample_per_issue_data_dict(
            issues, 
            is_final="TRUE",
            justifications=["Test justification"],
            short_justifications="Test summary"
        )
        
        tracker = TestUtils.create_sample_tracker(issues_dict=per_issue_data, config=self.mock_config)
        # Set realistic metrics that should work with ExcelWriter
        tracker.metrics = {
            "total_issues": 2,
            "predicted_true_positives": {"test_issue_1"},  # Set with 1 item
            "predicted_true_positives_count": 1,
            "predicted_false_positives": {"test_issue_2"},  # Set with 1 item
            "predicted_false_positives_count": 1,
            "has_ground_truth": True,
            "actual_true_positives": ["actual_tp_1"],  # List with 1 item
            "actual_true_positives_count": 1,
            "actual_false_positives": ["actual_fp_1"],  # List with 1 item
            "actual_false_positives_count": 1,
            "confusion_matrix": {
                "true_positives": 1,
                "true_negatives": 0,
                "false_positives": 1,
                "false_negatives": 0
            },
            "accuracy": 0.5,
            "precision": 0.5,
            "recall": 1.0,
            "f1_score": 0.67
        }
        
        # Create a temporary file for Excel output
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_excel_path = temp_file.name
        
        try:
            # Mock config attributes needed for ExcelWriter
            self.mock_config.OUTPUT_FILE_PATH = temp_excel_path
            self.mock_config.INPUT_REPORT_FILE_PATH = "test_report.html"  # Not a URL, so no Google Sheets
            self.mock_config.AGGREGATE_RESULTS_G_SHEET = None  # No aggregate sheet
            self.mock_config.RUN_WITH_CRITIQUE = False
            self.mock_config.SHOW_FINAL_JUDGE_CONTEXT = False
            
            with patch('sast_agent_workflow.tools.write_results.convert_tracker_to_summary_data') as mock_convert:
                # Create realistic summary_data that matches what ExcelWriter expects
                mock_summary_info_1 = Mock()
                mock_summary_info_1.llm_response.investigation_result = "TRUE POSITIVE"
                mock_summary_info_1.llm_response.short_justifications = "Test hint 1"
                mock_summary_info_1.llm_response.justifications = ["Test justification 1"]
                mock_summary_info_1.llm_response.recommendations = ["Test recommendation 1"]
                mock_summary_info_1.metrics = {"answer_relevancy": 0.8}
                mock_summary_info_1.critique_response = "Test critique"
                mock_summary_info_1.context = "Test context"
                
                mock_summary_info_2 = Mock()
                mock_summary_info_2.llm_response.investigation_result = "FALSE POSITIVE"
                mock_summary_info_2.llm_response.short_justifications = "Test hint 2"
                mock_summary_info_2.llm_response.justifications = ["Test justification 2"]
                mock_summary_info_2.llm_response.recommendations = ["Test recommendation 2"]
                mock_summary_info_2.metrics = {"answer_relevancy": 0.6}
                mock_summary_info_2.critique_response = "Test critique 2"
                mock_summary_info_2.context = "Test context 2"
                
                mock_convert.return_value = [
                    (issues[0], mock_summary_info_1),
                    (issues[1], mock_summary_info_2)
                ]
                
                # testing - Call the real write_results function (no mocking of write_to_excel_file)
                result_tracker = await TestUtils.run_single_fn(write_results, self.write_results_config, self.builder, tracker)
                
                # assertion
                self.assertIsInstance(result_tracker, SASTWorkflowTracker)
                
                # Verify the Excel file was actually created
                self.assertTrue(os.path.exists(temp_excel_path))
                self.assertGreater(os.path.getsize(temp_excel_path), 0)  # File is not empty
                
                # Additional validation: try to read the Excel file to ensure it's valid
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(temp_excel_path)
                    
                    # Verify expected worksheets exist
                    self.assertIn("AI report", workbook.sheetnames)
                    self.assertIn("Confusion Matrix", workbook.sheetnames)
                    
                    # Verify AI report worksheet has data
                    ai_report = workbook["AI report"]
                    self.assertIsNotNone(ai_report["A1"].value)  # Should have headers
                    self.assertIsNotNone(ai_report["A2"].value)  # Should have data
                    
                    # Verify Confusion Matrix worksheet has data
                    confusion_matrix = workbook["Confusion Matrix"]
                    self.assertIsNotNone(confusion_matrix["A1"].value)  # Should have content
                    
                    workbook.close()
                    
                except ImportError:
                    # If openpyxl is not available, just verify file exists and has content
                    self.assertGreater(os.path.getsize(temp_excel_path), 1000)  # Reasonable file size
                    
        finally:
            # Clean up temporary file
            if os.path.exists(temp_excel_path):
                os.unlink(temp_excel_path)


if __name__ == '__main__':
    unittest.main() 